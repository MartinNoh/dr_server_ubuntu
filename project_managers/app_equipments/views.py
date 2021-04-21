import os, datetime
from django.db.models import Q
from django.shortcuts import redirect, render, get_object_or_404
from .models import Usage, User, Device
from .forms import UsageEditForm, DeviceNewForm
import openpyxl
from django.http import HttpResponse, Http404


# home.html 페이지를 부르는 index 함수
def home(request):
    return render(request, 'app_equipments/base/home.html')


def check_total(request):

    info_list = []
    device_list = Device.objects.all()

    search_key = request.GET.get('search_key')  # 검색어 가져오기
    #print(search_key)
    if search_key:  # 만약 검색어가 존재하면
        device_list = device_list.filter(Q(category__icontains=search_key) | Q(sort__icontains=search_key) | Q(spec__icontains=search_key) | Q(is_assets__icontains=search_key)).distinct()  # 해당 검색어를 포함한 queryset 가져오기

    for i in device_list:
        usage_amount = Usage.objects.filter(device_id=i.device_id)

        info_dic = {
         'device_id': i.device_id,
         'category': i.category,
         'sort': i.sort,
         'purchase_date': i.purchase_date,
         'spec': i.spec,
         'is_assets': i.is_assets,
         'etc': i.etc,
         'total': i.amount,
         'amounts': len(usage_amount),
         'remains': i.amount - len(usage_amount),
        }
        #print(info_dic)

        info_list.append(info_dic)
        #print(info_list)

    return render(request, 'app_equipments/menu/check_total.html', {'info_list': info_list,})


def check_total_new(request):

    # new device
    if request.method == "POST":
        form = DeviceNewForm(request.POST)
        if form.is_valid():
            new_device = form.save()

        return redirect('check_total')
    else:
        form = DeviceNewForm()

    return render(request, 'app_equipments/menu/check_total_new.html', {'form': form, })


def check_total_update(request, device_id):
    device = Device.objects.get(device_id=device_id)
    if request.method == "POST":
        form = DeviceNewForm(request.POST)
        if form.is_valid():
            device.category = form.cleaned_data['category']
            device.sort = form.cleaned_data['sort']
            device.spec = form.cleaned_data['spec']
            device.amount = form.cleaned_data['amount']
            device.purchase_date = form.cleaned_data['purchase_date']
            device.is_assets = form.cleaned_data['is_assets']
            device.etc = form.cleaned_data['etc']
            device.save()

        return redirect('check_total')
    else:
        form = DeviceNewForm(instance=device)

    return render(request, 'app_equipments/menu/check_total_update.html', {'form': form, 'device_id': device_id,})


def check_total_delete(request, device_id):
    device = Device.objects.filter(device_id=device_id)
    device[0].delete()

    return redirect('/check_total/')


def download_excel(request):
    # 장비 정보를 리스트 담기
    device_id = []
    category = []
    sort = []
    purchase_date = []
    spec = []
    is_assets = []
    etc = []
    total = []
    amounts = []
    remains = []
    device_list = Device.objects.all()
    for i in device_list:
        device_id.append(i.device_id)
        category.append(i.category)
        sort.append(i.sort)
        purchase_date.append(i.purchase_date)
        spec.append(i.spec)
        is_assets.append(i.is_assets)
        etc.append(i.etc)
        total.append(i.amount)

        usage_amount = Usage.objects.filter(device_id=i.device_id)
        amounts.append(len(usage_amount))
        remains.append(i.amount - len(usage_amount))

    user_seat = []
    user_name = []
    user_category = []
    user_sort = []
    user_purchase_date = []
    user_spec = []
    user_is_assets = []
    user_etc = []
    # 직원 정보를 리스트 담기
    user_list = User.objects.all()
    #print(user_list)
    for i in user_list:
        usage_list = Usage.objects.filter(user_id=i.user_id)
        #print(usage_list)
        for j in usage_list:
            split_list = str(j).split('|')
            #print(split_list)
            user_seat.append(split_list[0].strip())
            user_name.append(split_list[1].strip())

            user_category.append(split_list[2].strip())
            user_sort.append(split_list[3].strip())
            user_purchase_date.append(split_list[4].strip())
            user_device = Device.objects.filter(category=split_list[2].strip(), sort=split_list[3].strip(),
                                                purchase_date=split_list[4].strip())
            user_spec.append(user_device[0].spec)
            user_is_assets.append(user_device[0].is_assets)
            user_etc.append(user_device[0].etc)

    users_folder_path = os.path.expanduser('~')
    downloads_path = os.path.join(users_folder_path, 'Downloads', 'GJAC_Equipments')
    file_name = 'GJAC_Equipments_' + datetime.datetime.now().strftime('%Y%m%d') + '.xlsx'
    output_path = os.path.join(downloads_path, file_name)
    try:
        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path)
    except OSError:
        print('Error: Creating directory. ' + downloads_path)


    # 서버 다운로드 파일 백업
    wb = openpyxl.Workbook()

    sheet1 = wb.active
    sheet1.title = "Total"
    sheet2 = wb.create_sheet("Usage")
    #sheet3 = wb.create_sheet("DB_User")
    #sheet4 = wb.create_sheet("DB_Device")
    #sheet5 = wb.create_sheet("DB_Usage")

    subject1 = ["총량", "사용량", "잔여량", "범주", "종류", "스펙", "구매일자", "자산여부", "기타"]
    sheet1.append(subject1)
    for i in range(len(device_id)):
        total_value = [
            int(total[i]), int(amounts[i]), int(remains[i]),
            str(category[i]), str(sort[i]), str(spec[i]),
            str(purchase_date[i]), str(is_assets[i]),  str(etc[i])
        ]
        sheet1.append(total_value)

    subject2 = ["위치번호", "이름", "범주", "종류", "스펙", "구매일자", "자산여부", "기타"]
    sheet2.append(subject2)
    for i in range(len(user_seat)):
        usage_value = [
            str(user_seat[i]), str(user_name[i]), str(user_category[i]), str(user_sort[i]),
            str(user_spec[i]), str(user_purchase_date[i]), str(user_is_assets[i]), str(user_etc[i])
        ]
        sheet2.append(usage_value)

    #subject3 = ["user_id", "name", "rank", "seat"]
    #sheet3.append(subject3)

    #subject4 = ["device_id", "purchase_date", "category", "sort", "amount", "spec", "is_assets", "etc"]
    #sheet4.append(subject4)

    #subject5 = ["usage_id", "user_id", "device_id", "updated_at"]
    #sheet5.append(subject5)


    wb.save(output_path)

    try:
        # 클라이언트 다운로드 rb 파일모드
        with open(output_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename=' + file_name

            return response
    except:
        return render(request, 'app_equipments/menu/download_excel.html', {})


def check_seat(request, office, seat):

    # get seat list
    seat_all = User.objects.all()

    # get user info
    user = User.objects.get(seat=seat)
    #print('3 :', user)

    # get usage info
    usage = Usage.objects.filter(user_id=user.user_id).order_by('device_id')
    #print('4 :', usage)

    # make usage list
    device_usage_info = []
    for i in usage:
        device = str(i.device_id).split('|')
        #print('5 :', device)
        #print('6 :', device[0])
        #print('7 :', device[1])
        device_spec = Device.objects.filter(category=device[0].strip(), sort=device[1].strip())
        #print('8 :', device_usage)
        #print('9 :', device_usage[0].spec)
        device_info = {
            'device_id': device_spec[0].device_id,
            'category': device[0],
            'sort': device[1],
            'spec': device_spec[0].spec,
            'is_assets': device_spec[0].is_assets
        }
        device_usage_info.append(device_info)
        #print('10 :', device_usage_info)

    # new usage
    if request.method == "POST":
        form = UsageEditForm(request.POST)
        if form.is_valid():
            new_usage = form.save(commit=False)
            new_usage.user_id = user
            new_usage.save()

            seat = user.seat
            get_page = '/check_seat/' + str(office) + '/' + str(seat)

        return redirect(get_page)
    else:
        form = UsageEditForm()

    return render(request, 'app_equipments/menu/check_seat.html', {'seat_all': seat_all, 'user': user, 'device_usage_info': device_usage_info, 'form': form, 'office': office})


def check_seat_delete(request, user_id, device_id, office):
    usage = Usage.objects.filter(user_id=user_id, device_id=device_id)
    #print(usage)
    if len(usage) > 0:
        usage[0].delete()

    user = User.objects.get(user_id=user_id)
    seat = user.seat
    get_page = '/check_seat/' + str(office) + '/' + str(seat)

    return redirect(get_page)


def notification(request):
    return render(request, 'app_equipments/menu/notification.html')